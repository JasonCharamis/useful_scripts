import pandas as pd
import argparse
import re
import openpyxl
import os

parser = argparse.ArgumentParser()
parser.add_argument('--xlsx', type=str, help='Input xlsx spreadsheet file.')
args = parser.parse_args()

def xlsx_to_tsv(xlsx_file):
    """
    Convert an Excel XLSX file to TSV file(s)
    """
    # Load the XLSX file
    workbook = openpyxl.load_workbook(xlsx_file)
    
    # Get the base name of the xlsx file (without extension)
    base_name = os.path.splitext(xlsx_file)[0]
    
    # Process each worksheet
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Create a list of rows
        rows = []
        for row in worksheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(str(cell.value) if cell.value is not None else '')
            rows.append('\t'.join(row_data))
        
        # Determine the output file name
        if len(workbook.sheetnames) == 1:
            tsv_file = f"{base_name}.tsv"
        else:
            tsv_file = f"{base_name}_{sheet_name}.tsv"
        
        # Write the TSV file
        with open(tsv_file, 'w', encoding='utf-8') as tsv_file:
            tsv_file.write('\n'.join(rows))
        
        print(f"Created: {tsv_file}")

if __name__ == "__main__":
    xlsx_to_tsv(xlsx_file=args.xlsx)
